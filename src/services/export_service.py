"""Data export service for multiple formats."""

import json
import csv
from typing import Dict, Any, List
from io import StringIO, BytesIO


class ExportService:
    """Service for exporting processed data in multiple formats."""

    def export_data(
        self, data: Dict[str, Any], format: str, options: Dict[str, Any] = None
    ) -> bytes:
        if options is None:
            options = {}
        """Export data in the requested format.

        Args:
            data: Processed extraction data
            format: Export format ('csv', 'json', 'pdf', 'excel', 'xml', 'html')
            options: Export options

        Returns:
            Exported data as bytes
        """
        if options is None:
            options = {}

        exporters = {
            "csv": self._export_csv,
            "json": self._export_json,
            "excel": self._export_excel,
            "xml": self._export_xml,
            "html": self._export_html,
        }

        if format not in exporters:
            raise ValueError(f"Unsupported export format: {format}")

        return exporters[format](data, options)

    def _export_csv(self, data: Dict[str, Any], options: Dict[str, Any]) -> bytes:
        """Export data as CSV."""
        output = StringIO()

        # Export each table
        for i, table in enumerate(data.get("tables", [])):
            if i > 0:
                output.write("\n")  # Empty line between tables

            writer = csv.writer(output)

            # Table name
            writer.writerow([f"Table: {table.get('name', f'Table_{i + 1}')}"])
            writer.writerow([])

            # Headers
            columns = table.get("columns", [])
            if columns:
                writer.writerow(columns)

                # Data rows
                rows = table.get("rows", [])
                for row in rows:
                    writer.writerow([row.get(col, "") for col in columns])

        return output.getvalue().encode("utf-8")

    def _export_json(self, data: Dict[str, Any], options: Dict[str, Any]) -> bytes:
        """Export data as JSON."""
        # Clean data for JSON export
        clean_data = {
            "metadata": data.get("metadata", {}),
            "tables": data.get("tables", []),
            "text": data.get("text", ""),
            "export_timestamp": options.get("timestamp", None),
        }

        return json.dumps(clean_data, indent=2, ensure_ascii=False).encode("utf-8")

    def _export_excel(self, data: Dict[str, Any], options: Dict[str, Any]) -> bytes:
        """Export data as Excel (.xlsx)."""
        try:
            from openpyxl import Workbook

            wb = Workbook()

            # Create sheets for each table
            for i, table in enumerate(data.get("tables", [])):
                sheet_name = table.get("name", f"Table_{i + 1}")[:31]  # Excel limit
                ws = wb.create_sheet(title=sheet_name)

                # Headers
                columns = table.get("columns", [])
                if columns:
                    for col_num, header in enumerate(columns, 1):
                        ws.cell(row=1, column=col_num, value=header)

                    # Data rows
                    rows = table.get("rows", [])
                    for row_num, row_data in enumerate(rows, 2):
                        for col_num, header in enumerate(columns, 1):
                            value = row_data.get(header, "")
                            ws.cell(row=row_num, column=col_num, value=value)

            # Remove default sheet if we have tables
            if data.get("tables"):
                wb.remove(wb["Sheet"])

            # Save to bytes
            buffer = BytesIO()
            wb.save(buffer)
            return buffer.getvalue()

        except ImportError:
            raise Exception("openpyxl not installed. Install with: pip install openpyxl")

    def _export_xml(self, data: Dict[str, Any], options: Dict[str, Any]) -> bytes:
        """Export data as XML."""

        def dict_to_xml(data_dict: Dict[str, Any], root_name: str = "data") -> str:
            """Convert dictionary to XML string."""
            xml_parts = [f"<{root_name}>"]

            for key, value in data_dict.items():
                if isinstance(value, list):
                    for i, item in enumerate(value):
                        if isinstance(item, dict):
                            xml_parts.append(dict_to_xml(item, f"{key}_{i}"))
                        else:
                            xml_parts.append(f"<{key}>{str(item)}</{key}>")
                elif isinstance(value, dict):
                    xml_parts.append(dict_to_xml(value, key))
                else:
                    xml_parts.append(f"<{key}>{str(value)}</{key}>")

            xml_parts.append(f"</{root_name}>")
            return "".join(xml_parts)

        xml_content = '<?xml version="1.0" encoding="UTF-8"?>\n'
        xml_content += dict_to_xml(data)

        return xml_content.encode("utf-8")

    def _export_html(self, data: Dict[str, Any], options: Dict[str, Any]) -> bytes:
        """Export data as HTML."""
        html_parts = [
            "<!DOCTYPE html>",
            '<html lang="en">',
            "<head>",
            '<meta charset="UTF-8">',
            '<meta name="viewport" content="width=device-width, initial-scale=1.0">',
            "<title>Data Export</title>",
            "<style>",
            "body { font-family: Arial, sans-serif; margin: 20px; }",
            "table { border-collapse: collapse; width: 100%; margin-bottom: 20px; }",
            "th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
            "th { background-color: #f2f2f2; }",
            "tr:nth-child(even) { background-color: #f9f9f9; }",
            "h2 { color: #333; }",
            "h3 { color: #666; }",
            "</style>",
            "</head>",
            "<body>",
            "<h1>Data Export</h1>",
        ]

        # Add metadata
        if data.get("metadata"):
            html_parts.append("<h2>Metadata</h2>")
            html_parts.append("<table>")
            for key, value in data["metadata"].items():
                html_parts.append(f"<tr><td>{key}</td><td>{value}</td></tr>")
            html_parts.append("</table>")

        # Add tables
        for i, table in enumerate(data.get("tables", [])):
            html_parts.append(f"<h2>Table: {table.get('name', f'Table_{i + 1}')}</h2>")
            html_parts.append("<table>")

            # Headers
            columns = table.get("columns", [])
            if columns:
                html_parts.append("<tr>")
                for col in columns:
                    html_parts.append(f"<th>{col}</th>")
                html_parts.append("</tr>")

                # Data rows
                rows = table.get("rows", [])
                for row in rows:
                    html_parts.append("<tr>")
                    for col in columns:
                        value = row.get(col, "")
                        html_parts.append(f"<td>{value}</td>")
                    html_parts.append("</tr>")

            html_parts.append("</table>")

        # Add text content if available
        if data.get("text"):
            html_parts.append("<h2>Text Content</h2>")
            html_parts.append("<pre>")
            html_parts.append(data["text"])
            html_parts.append("</pre>")

        html_parts.extend(["</body>", "</html>"])

        return "\n".join(html_parts).encode("utf-8")
