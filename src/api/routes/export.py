"""Export API endpoints for data export in multiple formats."""

import json
import csv
from typing import Dict, Any
from io import StringIO, BytesIO

from fastapi import APIRouter, HTTPException, status, Depends
from fastapi.responses import StreamingResponse, JSONResponse

from src.dependencies import CurrentUser
from src.services.extraction_service import ExtractionService

router = APIRouter(prefix="/export", tags=["export"])


@router.post("/csv")
async def export_csv(data: Dict[str, Any], user: CurrentUser) -> StreamingResponse:
    """Export data as CSV file.

    Handles both tabular data and text-only data.
    """
    try:
        # Create CSV content
        output = StringIO()
        writer = csv.writer(output)

        # Check if we have tabular data
        if data.get("tables") and len(data["tables"]) > 0:
            # Use first table for CSV export
            table = data["tables"][0]

            # Write headers
            writer.writerow(table["columns"])

            # Write data rows
            for row in table["rows"]:
                csv_row = [str(row.get(col, "")) for col in table["columns"]]
                writer.writerow(csv_row)
        else:
            # Handle text-only data - create a simple single-column table
            writer.writerow(["Content"])

            # Add text content if available
            if data.get("text"):
                # Split text into lines and create rows
                lines = data["text"].split("\n")
                for line in lines:
                    if line.strip():  # Skip empty lines
                        writer.writerow([line.strip()])

            # If no text either, add a placeholder
            if not data.get("text"):
                writer.writerow(["No content available"])

        csv_content = output.getvalue()
        output.close()

        # Create streaming response
        def generate():
            yield csv_content

        # Set filename based on available data
        if data.get("tables") and len(data["tables"]) > 0:
            filename = f"export_{data['tables'][0].get('name', 'data')}.csv"
        else:
            filename = "export_text.csv"

        return StreamingResponse(
            generate(),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"CSV export failed: {str(e)}"
        )


@router.post("/json")
async def export_json(data: Dict[str, Any], user: CurrentUser) -> StreamingResponse:
    """Export data as JSON file."""
    try:
        json_content = json.dumps(data, indent=2, ensure_ascii=False)

        def generate():
            yield json_content

        return StreamingResponse(
            generate(),
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=export.json"},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"JSON export failed: {str(e)}",
        )


@router.post("/excel")
async def export_excel(data: Dict[str, Any], user: CurrentUser) -> StreamingResponse:
    """Export data as Excel file (.xlsx)."""
    try:
        # Create Excel file using openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        # Check if we have tabular data
        if data.get("tables") and len(data["tables"]) > 0:
            # Process each table
            for table_idx, table in enumerate(data["tables"]):
                if table_idx > 0:
                    # Create new sheet for additional tables
                    ws = wb.create_sheet(title=f"Table {table_idx + 1}")

                # Write headers with styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )

                for col_idx, col_name in enumerate(table["columns"], 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill

                # Write data rows
                for row_idx, row_data in enumerate(table["rows"], 2):
                    for col_idx, col_name in enumerate(table["columns"], 1):
                        value = row_data.get(col_name, "")
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # Auto-adjust column widths
                for col_idx, col_name in enumerate(table["columns"], 1):
                    column_letter = ws.cell(row=1, column=col_idx).column_letter
                    max_length = (
                        max(len(str(row_data.get(col_name, ""))) for row_data in table["rows"])
                        if table["rows"]
                        else 10
                    )
                    max_length = max(max_length, len(col_name))
                    ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        else:
            # Handle text-only data - create a simple single-column sheet
            ws.cell(row=1, column=1, value="Content")

            # Add text content if available
            if data.get("text"):
                # Split text into lines and create rows
                lines = data["text"].split("\n")
                for row_idx, line in enumerate(lines, 2):
                    if line.strip():  # Skip empty lines
                        ws.cell(row=row_idx, column=1, value=line.strip())

            # Auto-adjust column width
            ws.column_dimensions["A"].width = 50

        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        def generate():
            yield excel_buffer.getvalue()

        filename = "export.xlsx"

        return StreamingResponse(
            generate(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export requires openpyxl. Install with: pip install openpyxl",
        )
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel export failed: {str(e)}",
        )


@router.post("/xml")
async def export_xml(data: Dict[str, Any], user: CurrentUser) -> StreamingResponse:
    """Export data as XML file."""
    try:

        def dict_to_xml(data_dict, root_name="data"):
            """Convert dictionary to XML string."""

            def _convert(obj, name):
                if isinstance(obj, dict):
                    xml_parts = [f"<{name}>"]
                    for key, value in obj.items():
                        xml_parts.append(_convert(value, key))
                    xml_parts.append(f"</{name}>")
                    return "".join(xml_parts)
                elif isinstance(obj, list):
                    xml_parts = []
                    for item in obj:
                        xml_parts.append(_convert(item, "item"))
                    return "".join(xml_parts)
                else:
                    # Escape XML characters
                    value_str = (
                        str(obj).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    )
                    return f"<{name}>{value_str}</{name}>"

            return f'<?xml version="1.0" encoding="UTF-8"?>\n{_convert(data_dict, root_name)}'

        xml_content = dict_to_xml(data)

        def generate():
            yield xml_content

        return StreamingResponse(
            generate(),
            media_type="application/xml",
            headers={"Content-Disposition": "attachment; filename=export.xml"},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"XML export failed: {str(e)}"
        )


@router.post("/html")
async def export_html(data: Dict[str, Any], user: CurrentUser) -> StreamingResponse:
    """Export data as HTML file with table formatting."""
    try:
        html_parts = [
            "<!DOCTYPE html>",
            '<html lang="en">',
            "<head>",
            '    <meta charset="UTF-8">',
            '    <meta name="viewport" content="width=device-width, initial-scale=1.0">',
            "    <title>Extracted Data</title>",
            "    <style>",
            "        body { font-family: Arial, sans-serif; margin: 20px; }",
            "        h1 { color: #333; }",
            "        table { border-collapse: collapse; width: 100%; margin: 20px 0; }",
            "        th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }",
            "        th { background-color: #f2f2f2; font-weight: bold; }",
            "        tr:nth-child(even) { background-color: #f9f9f9; }",
            "        .metadata { background-color: #e8f4f8; padding: 15px; border-radius: 5px; margin: 20px 0; }",
            "        .text-content { background-color: #f8f8f8; padding: 15px; border-radius: 5px; margin: 20px 0; white-space: pre-wrap; }",
            "    </style>",
            "</head>",
            "<body>",
            "    <h1>Extracted Data Report</h1>",
        ]

        # Add metadata if available
        if data.get("metadata"):
            html_parts.append('    <div class="metadata">')
            html_parts.append("        <h2>Document Metadata</h2>")
            html_parts.append("        <ul>")
            for key, value in data["metadata"].items():
                html_parts.append(f"            <li><strong>{key}:</strong> {value}</li>")
            html_parts.append("        </ul>")
            html_parts.append("    </div>")

        # Add extracted text if available
        if data.get("text"):
            html_parts.append('    <div class="text-content">')
            html_parts.append("        <h2>Extracted Text</h2>")
            html_parts.append(f"        <p>{data['text'].replace(chr(10), '<br>')}</p>")
            html_parts.append("    </div>")

        # Add tables
        if data.get("tables"):
            for table_idx, table in enumerate(data["tables"]):
                html_parts.append(
                    f"    <h2>Table {table_idx + 1}: {table.get('name', 'Data')}</h2>"
                )
                html_parts.append("    <table>")
                html_parts.append("        <thead>")
                html_parts.append("            <tr>")

                # Table headers
                for col in table["columns"]:
                    html_parts.append(f"                <th>{col}</th>")
                html_parts.append("            </tr>")
                html_parts.append("        </thead>")
                html_parts.append("        <tbody>")

                # Table rows
                for row in table["rows"]:
                    html_parts.append("            <tr>")
                    for col in table["columns"]:
                        value = row.get(col, "")
                        html_parts.append(f"                <td>{value}</td>")
                    html_parts.append("            </tr>")

                html_parts.append("        </tbody>")
                html_parts.append("    </table>")

        # Add note if available
        if data.get("note"):
            html_parts.append('    <div class="metadata">')
            html_parts.append(f"        <p><strong>Note:</strong> {data['note']}</p>")
            html_parts.append("    </div>")

        html_parts.extend(["</body>", "</html>"])

        html_content = "\n".join(html_parts)

        def generate():
            yield html_content

        return StreamingResponse(
            generate(),
            media_type="text/html",
            headers={"Content-Disposition": "attachment; filename=export.html"},
        )

    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"HTML export failed: {str(e)}",
        )
